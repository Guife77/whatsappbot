import webbrowser
import time
import urllib.parse
import pyautogui
import pymysql.cursors
import openpyxl

# Constants
WHATSAPP_WEB_URL = 'https://web.whatsapp.com/send'
ERROR_LOG_FILE = ''
TELEFONES_ENVIADOS_FILE = ''
DATABASE_HOST = ''
DATABASE_USER = ''
DATABASE_PASSWORD = '@'
DATABASE_NAME = ''

# Funções
def read_numbers_from_file(file_path):
    """Read phone numbers from an Excel file"""
    try:
        wb = openpyxl.load_workbook(file_path)
        sheet = wb['Sheet1']
        numbers = [row[0] for row in sheet.iter_rows(min_row=1, values_only=True) if row[0]]
        return numbers
    except Exception as e:
        print(f'Error reading Excel file: {e}')
        return []

def write_to_excel(file_path, phone_number):
    """Write a phone number to an Excel file"""
    try:
        wb = openpyxl.load_workbook(file_path)
        sheet = wb['Sheet1']
        for row in range(2, sheet.max_row + 2):
            if sheet[f'A{row}'].value is None:
                sheet[f'A{row}'] = phone_number
                break
        wb.save(file_path)
    except Exception as e:
        print(f'Error writing to Excel file: {e}')

def send_whatsapp_message(phone_number, message):
    try:
        link = f'{WHATSAPP_WEB_URL}?phone={phone_number}&text={urllib.parse.quote(message)}'
        webbrowser.open(link)
        time.sleep(10)
        pyautogui.press('enter')
        time.sleep(6)
        pyautogui.hotkey('ctrl', 'w')
        time.sleep(1)
    except Exception as e:
        print(f'Error sending WhatsApp message: {e}')

def log_error(name, phone_number):
    with open(ERROR_LOG_FILE, 'a', newline='', encoding='utf-8') as file:
        file.write(f'{name},{phone_number}\n')

def main():

    con = pymysql.connect(
        host=DATABASE_HOST,
        user=DATABASE_USER,
        password=DATABASE_PASSWORD,
        database=DATABASE_NAME,
        cursorclass=pymysql.cursors.DictCursor
    )

    phone_numbers = read_numbers_from_file(TELEFONES_ENVIADOS_FILE)

    last_processed_id = 0

    while True:
        with con.cursor() as cursor:
            query = f"SELECT nome, telefone, pagamento_final, lote FROM ??? WHERE Id > {last_processed_id} ORDER BY Id ASC"
            cursor.execute(query)
            results = cursor.fetchall()

        for result in results:
            name = result['nome']
            lote = result['lote']
            phone_number = result['telefone']
            pago = result ['pagamento_final']
            link1 = ''
            
            if pago == "vazio":
                continue

            message = f"""  Olá {name},

                Parabéns! Sua inscrição para o EDOM 10ª Edição foi concluída com sucesso. Estamos muito animados em tê-lo(a) conosco nesse evento tão especial!

                Agora, estamos cada vez mais próximos do grande dia, e para garantir ainda mais conforto e organização,
                convidamos você a selecionar seu assento para o evento. Basta acessar o link abaixo e escolher a sua cadeira:

                Relembrando que seu lote é o {lote} seja conciente e escolha a cadeira correspondente ao seu lote.

                {link1}

                Estamos ansiosos para encontrá-lo(a) no EDOM 10ª Edição!

                Atenciosamente,
                nome"""

            if phone_number in phone_numbers:
                    print("Já foi enviado!")
                    continue

            try:
                    send_whatsapp_message(phone_number, message)
                    phone_numbers.append(phone_number)
                    write_to_excel(TELEFONES_ENVIADOS_FILE, phone_number)
            except Exception as e:
                    log_error(name, phone_number)
                    print(f'Erro em mandar mensagem para {name}: {e}')

        time.sleep(15)

if __name__ == '__main__':
    main()
